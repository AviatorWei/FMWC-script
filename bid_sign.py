#北大未名BBS体育游戏版FM-America 2019游戏
#BY survivor@BDWM
#Python 3.7

import re
import pickle
import string
import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import chain

###基本信息

# 2022世界杯参赛国家，共32个
Nations = ['Ecuador', 'Netherlands', 'Qatar', 'Senegal', 'England', 'Iran', 'USA', 'Wales',
           'Argentina', 'Mexico', 'Poland', 'SaudiArabia', 'Australia', 'Denmark', 'France', 'Tunisia',
           'CostaRica', 'Germany', 'Japan', 'Spain', 'Belgium', 'Canada', 'Croatia', 'Morocco',
           'Brazil', 'Cameroon', 'Serbia', 'Switzerland', 'Ghana', 'Portugal', 'SouthKorea', 'Uruguay']

# FMWC游戏参赛球队，共24支
Participants = [
    ['Sang', 'ESP'], ['angelqi', 'BRA'], ['augustusc', 'MEX'], ['aidengazhaer', 'NED'],
    ['since', 'QAT'], ['KakaHiguain', 'MAR'], ['Nocchiere', 'DEN'], ['sixingdeguo', 'SRB'],
    ['Arjen', 'SUI'], ['Zeymax', 'ARG'], ['RealMadrid', 'POR'], ['rhyshm', 'FRA'],
    ['HydraliskIII', 'KSA'], ['solojuve', 'ECU'], ['Montella', 'CAN'], ['pkuarsene', 'GHA'],
    ['weilovebvb', 'CRC'], ['twa', 'WAL'], ['linsage', 'SEN'], ['IanWalls', 'ENG'],
    ['jiamingpku', 'URU'], ['Augustus', ''], ['Jimmywiki', 'JPN'], ['survivor', 'USA'],
]

Teams = [x[1] for x in Participants]
# 各队主教练ID，顺序与Teams对应。
Managers = [x[0] for x in Participants]

Budget0 = 600 #各队初始资金数额
Orders1 = [4,10,12,7,11,9,1,8,3,6,5,2] #首轮暗标各队投标顺序


def dictize(keys, values): #输入键列表keys和值列表values，输出对应的字典dic。
    assert len(keys) == len(values), f"Length not equal {keys}, {values}"
    dic = {}
    for key, value in zip(keys, values):
        assert key not in dic, f"Key existed {key}-{value}, original {dic[key]}"
        dic[key] = value
    return dic


ManagersDic = dictize(Teams, Managers) #玩家信息的字典
Budgets0Dic = dictize(Teams, [Budget0]*len(Teams)) #资金信息的字典
Orders1Dic = dictize(Teams, Orders1) 
Black1Dic = dictize(Teams, [[]]*len(Teams))


def str_len(string):#输入字符串string，输出string的字符个数，适配中文。
    try:
        row_l=len(string)
        utf8_l=len(string.encode('utf-8'))
        return int((utf8_l-row_l)/2+row_l)
    except:
        return None


def pos_value(p): #输入位置的单个字符，输出数值0，1，2，3.
    positions = ['G', 'D', 'M', 'F']
    pos_dict = dict([(v, i) for i, v in enumerate(positions)])
    return pos_dict[p]


def SquadToQuad(List): #输入一个多名球员阵容List，输出其四个位置上球员个数的列表。
    output = [0]*4
    for entry in List:
        pos = entry[2]
        output[pos_value(pos)] += 1
    return output


### 建立大名单的数据库
Roster1 = load_workbook('../FMWC2022大名单.xlsx')
Worksheet1 = Roster1['WC2022']


def BuildDatabase(worksheet):
    output = dict()
    for row in worksheet.iter_rows(min_row=2, max_col=4):
        entry = [x.value for x in row]
        key = tuple([entry[2], str(entry[3])])
        value = dict()
        value['name'] = entry[0].strip(' ') #球员姓名
        value['position'] = entry[1].strip(' ') #球员位置
        value['current'] = [] #球员当前所属玩家列表
        value['history'] = [] #球员属于过的全部玩家列表
        output[key] = value
    return output

# 存储当前数据库
Database = BuildDatabase(Worksheet1)
with open('../FMWC-2022-Database1.pickle', 'w+') as handle:
    pickle.dump(Database, handle, protocol=pickle.HIGHEST_PROTOCOL)

# 从pickle文件加载已有的数据库
with open('../FMWC-2022-Database1.pickle', 'rb') as file:
    Database = pickle.load(file)

### 人数要求
SquadUB = 16 #阵容人数上限
PosUB = [2,4,6,4] #阵容各位置人数上限
LineupLB = 5 #首发人数下限
LineupPosUB = [1,2,3,2] #首发各位置人数上限


def PosQuad(bid): #输入列表形式的标书bid，输出标书中四个位置的球员个数的列表。
    output = [0, 0, 0, 0]
    for entry in bid:
        position = entry[1]
        output[pos_value(position)] += 1
    return output


def IsQuadUBGood(quad):#输入阵容四个位置人数列表quad，输出其是否满足阵容人数上限要求。
    if sum(quad) > SquadUB:
        return False
    elif quad[0] > PosUB[0]:
        return False
    else:
        for index in range(1,4):
            if sum(quad[index:]) > sum(PosUB[index:]):
                return False
        return True


def NeededPlayer(bid, currentQuad): #输入标书bid和该玩家阵容中四个位置已有的球员个数列表currentQuad，输出在暗标全部命中的情况下，该玩家阵容需要补充的球员总数。
    quad = [PosQuad(bid)[index] + currentQuad[index] for index in range(4)]
    lineup = min(1, quad[0]) + min(7, quad[1] + 5, quad[1] + quad[2] + min(2, quad[3])) #第一项是标书在门将位置对首发的最大贡献，第二项是标书在后卫、中场、前锋位置对首发的最大贡献之和。
    return max(0, 5 - lineup)


# 读取xlsx标书
def read_bid(root, name, round):
    path = root + '/' + name.upper() + str(round) + '.xlsx'
    bid = load_workbook(path)
    sheet = bid.active
    output = []
    for row in sheet.iter_rows(min_row=2, max_col=6):
        entry = [x.value for x in row]
        if entry[0] is not None and entry[1] is not None:
            key = tuple([entry[4], str(entry[5])]) #key=(球员国家, 号码)
            name = entry[2].strip(' ') #球员姓名
            pos = entry[3].strip(' ') #球员位置
            order = entry[0] #排名
            price = entry[1]  # 出价
            assert isinstance(order, int), f'{path} order not int: {order}'
            assert isinstance(price, int), f'{path} price not int: {price}'
            bid = [key, pos, order, price]
            output.append(bid)
    return output


### 标书文本处理
def BidToList(text, nations, black):#输入单个球队的标书文本text，球队信息nations，该队不能签约的球员信息black，输出列表形式的标书，每条投标的格式为[(国家, 号码), 位置, 顺位, 出价]
    output = []
    poss = ['G','D','M','F']
    pos = 0 #当前位置
    order = 1 #当前顺位
    while 1:
        line = text.readline()
        if not line: #文本读取完毕
            return output 
        elif line == '\n': #读入用于分隔的空行
            pos += 1
            order = 1
        else: #读入一个投标
            #以下替换掉冗余字符
            line = line.replace('\n','')
            line = line.rstrip('m')
            line = line.rstrip('M')
            line = line.replace('#','')
            line = line.replace('号','')
            entry = re.split('[\s]+',line)
            bid = [poss[pos], order, ''] #投标信息的原型
            digits = [] #捕捉数字
            nation = ''
            for term in entry:
                if term in nations: #捕捉国家
                    nation = term
                elif term.isdigit(): 
                    digits.append(term)
            if len(digits)<2:
                print(text)
            price = int(digits[-1]) #捕捉最后一个数字，默认是出价
            number = int(digits[-2]) #捕捉倒数第二个数字，默认是号码
            key = (nation, str(number)) #创建球员身份的键值对
            if not key in black: #可以签约的球员
                bid[2] = price
                bid = [key] + bid
                output.append(bid) #添加列表形式的投标信息到输出
                order += 1


def BidsDic(root, teams, nations, blackdic, rd): #输入玩家信息teams，球队信息nations，不能签约球员信息的字典blackdic，输出各队列表形式标书的字典。
    bids = dict()
    for team in teams:
        # black = blackdic[team]
        bid = read_bid(root, team, rd)
        bids[team] = bid
    return bids


root = './bids-1'
Bids1 = BidsDic(root, Teams, Nations, Black1Dic, 1) #一轮暗标的列表形式标书的字典


### 列表标书处理
def TopPlayer(bid): #输入列表形式的标书bid，输出该标书中出价最高的一个投标。出价相同时，再依次比较位置, 顺位。
    if bid == []: #如果输入为空，返回空列表。
        return []
    else: #如果输入不为空
        sortedlist = sorted(bid, key=lambda x: (-x[3], pos_value(x[1]), x[2])) #依次比较：-出价, 位置数值, 顺位，都是数值小的顺序优先。
        return sortedlist[0] #返回出价最高的投标


def Budget(bid): #输入列表形式的标书bid，输出标书总出价。
    return sum([entry[-1] for entry in bid])


#输入玩家Team，其列表形式的标书List，球员数据库dictionary，该玩家当前阵容currentQuad和当前资金budget，检查标书合法性并返回[列表形式的所有有效投标bid,无效投标的报错信息文本announcement]。
def CheckBid(team, oriBid, db, currentQuad, budget):
    bid = []
    announcement = ''
    for entry in oriBid:
        key = entry[0]
        position = entry[1]
        price = entry[3]
        if not key in db.keys(): #检查球员是否在数据库中
            announcement = announcement + '无此球员     - ' + str(entry)
        elif db[key]['current'] != []: #检查球员是否已经被签约
            announcement = announcement + '已被签约     - ' + str(entry)
        elif team in db[key]['history']: #检查球员是否曾经被team签约
            announcement = announcement + '无资格签约    - ' + str(entry)
        elif db[key]['position'] != position: #检查球员的位置是否正确
            announcement = announcement + '位置错误     - ' + str(entry)
        elif price < 10: #检查出价是否至少为10
            announcement = announcement + '出价小于10   - ' + str(entry)
        else: #正确则放进bid列表
            bid.append(entry)
    while len(bid) + sum(currentQuad) > SquadUB: #标书多于SquadUB人时，移除最高价球员直到剩SquadUB人
        top = TopPlayer(bid)
        announcement = announcement + '人数超额     - ' + str(entry)
        bid.remove(top)
    while Budget(bid) > budget: #标书总出价大于剩余资金budget时，移除最高价球员直到总出价不大于budget
        top = TopPlayer(bid)
        announcement = announcement + '预算超额     - ' + str(entry)
        bid.remove(top)
    while Budget(bid) + 10*NeededPlayer(bid, currentQuad) > budget: #标书总出价使得全中可能导致凑不齐首发阵容人数下限，移除最高价球员直到满足要求
        top = TopPlayer(bid)
        announcement = announcement + '首发人数不足 - ' + str(entry)
        bid.remove(top)

    #去掉无效投标后，更新顺位
    # if bid == []:
    #     return [bid, announcement]
    # else:
    #     output = []
    #     currentpos = bid[0][1] #标书首行的位置
    #     counter = 0
    #     for entry in bid:
    #         pos = entry[0]
    #         if pos != currentpos:
    #             counter = 1
    #             currentpos = pos
    #         else:
    #             counter = counter + 1
    #         output.append([entry[0], entry[1], counter, entry[3]])
    output = bid
    if announcement != '':
        print(team, announcement)
    return [output, announcement]


def CompleteBid(bid, team, order): #输入同一玩家列表形式的所有有效投标bid, 每条投标添加队名和投标顺序成为完整标书 --  [玩家, (国家, 号码), 位置, 顺位, 出价, 投标顺序]
    return [[team] + entry + [order] for entry in bid]


### 竞标过程

def BidCompare(entries): #输入对同一球员的多笔有效投标的列表entries，输出中标的投标。
    if len(entries) == 1:
        return entries[0]
    else:
        ordered = sorted(entries, key = lambda entry : (-entry[4], entry[3], entry[5])) # 依次按照出价（从大到小）, 顺位（从小到大）, 投标顺序（从小到大）进行排序。
        return ordered[0]

def BidResult(bids, teams, dictionary, budgets): #输入列表形式的所有有效投标bids, 玩家teams，数据库dictionary和预算字典budgets，比较全部有效标书的投标，得出中标情况。
    players = dictionary.keys() #所有的球员列表
    profiles = dict.fromkeys(players, []) #创建每个球员全部被投标信息的字典，球员是键，投标信息的列表是值。
    newdictionary = copy.deepcopy(dictionary)
    newbudgets = budgets.copy()
    output = dict()
    for team in teams:
        output[team] = []
    for bid in bids:
        key = bid[1]
        current = profiles[key].copy()
        current.append(bid)
        profiles[key] = current #更新字典profile的值
    for player in profiles:
        if profiles[player] == []: #如果球员player不出现在暗标中，则直接跳过。
            continue
        else: #如果球员player出现在暗标中
            successbid = BidCompare(profiles[player]) #得到对球员player中标的投标
            team = successbid[0]
            price = successbid[4]
            output[team].append(successbid) #添加到输出
            profile = dictionary[player] #当前数据库中球员player的信息
            profile['current'] = [team] #添加其当前所属球队的信息
            profile['history'].append(team) #添加其曾经所属球队的信息
            newdictionary[player] = profile #更新数据库中球员player的信息
            newbudgets[team] = newbudgets[team] - price #更新玩家team的资金信息
    for team in teams:
        output[team] = sorted(output[team], key = lambda entry : pos_value(entry[2]))
    return [output, newdictionary, newbudgets] #输出各队中标信息output，新数据库newdictionary和新资金信息newbudgets


BidResult1 = BidResult(list(chain(*CompleteBids1.values())), Teams, Database, Budgets0Dic)
Squads1 = BidResult1[0]
Database1 = BidResult1[1]
Budgets1 = BidResult1[2]

### 输出文本处理
def LineToTxt(parts):#输入单行格式的参数parts = [part1, part2, ...]，输出按照此格式的文本。其中每个part = [内容, 长度, 对齐方式]
    output = ''
    for index in range(len(parts)):
        part = parts[index]
        string = str(part[0])
        length = part[1]
        align = part[2]
        if align == 'r':
            output = output + ' '*(max(length - str_len(string),0)) + string
        elif index == len(parts) - 1:
            output = output + string
        else:
            output = output + string + ' '*(max(length - str_len(string),0))
    output = output + '\n'
    return output

def SquadToText(squad, database, budget, team, manager): #输入一个玩家team的阵容squad，以及其资金budget和大师manager，输出该玩家阵容按指定格式的文本output。
    size = len(squad) #阵容人数
    output = ''
    output = team + ' ' + str(size) + '人' + ' '*(33 - len(team) - len(str(size))) + '剩余资金' + ' '*(4 - len(str(budget))) + str(budget) + 'm\n'
    output = output + manager + '\n'
    for entry in squad:
        nation = entry[1][0]
        number = entry[1][1] + '号'
        name = database[entry[1]]['name']
        position = entry[2]
        price = str(entry[4]) + 'm'
        info = [team, name, nation, number, position, price]
        card = len(info)
        lengths = [5,14,12,8,4,6]
        aligns = ['l', 'l', 'l', 'r', 'r', 'r']
        parts = [[string,length,align] for string, length, align in zip(info,lengths,aligns)]
        output += LineToTxt(parts)
    output += '\n'
    return output

def SquadsOutput(squads, database, budgets, teams, managers, filename): #输入各玩家的字典teams，其阵容的字典squads，资金的字典budgets，大师的字典managers，以及输出文件的文件名filename，输出各玩家阵容按指定格式的文本output并保存到filename.txt
    output = ''
    for key in teams:
        squad = squads[key]
        budget = budgets[key]
        manager = managers[key]
        output += SquadToText(squad, database, budget, key, manager)
    file = open(filename + '.txt','w')
    file.write(output)
    file.close()
    return output

SquadsOutput(Squads1, Database, Budgets1, Teams, ManagersDic, '一轮暗标后阵容')

###二轮暗标

Roster2 = load_workbook('-2019大名单2.xlsx') #二轮大名单

def UpdateDB(workbook, title, database): #输入excel工作簿workbook，工作表名title和旧数据库database，输出新数据库。
    output = copy.deepcopy(database)
    worksheet = workbook[title]
    for row in worksheet.iter_rows(min_row=2, max_col=5):
        entry = [x.value for x in row]
        key = tuple([entry[0], str(entry[1])])
        if not key in database.keys():
            value = dict()
            value['name'] = entry[2].strip(' ') #球员姓名
            value['position'] = entry[3].strip(' ') #球员位置
            value['current'] = [] #球员当前所属玩家列表
            value['history'] = [] #球员属于过的全部玩家列表
            output[key] = value
        elif database[key]['current'] != []:
            team = database[key]['current'][0]
            cell = row[4]
            #print(row)
            #cell.value = team
    #workbook.save(r'-2019大名单2.xlsx')
    return output

Database2 = UpdateDB(Roster2,'大名单', Database1) #次轮暗标数据库

'''
file = open('-2019-Database2.pickle','wb')
pickle.dump(Database2, file)
file.close()
'''

Orders2 = [3,9,11,7,6,2,8,10,4,1,12,5] #次轮暗标各队投标顺序
Orders2Dic = dictize(Teams, Orders2)

Bids2 = BidsDic(Teams, Nations, Black1Dic, 2) #二轮暗标的列表形式标书的字典
ValidBids2 =dict([(key, CheckBid(key, value, Database2, SquadToQuad(Squads1[key]), Budgets1[key])[0]) for key,value in Bids2.items()]) #二轮有效暗标的字典
CompleteBids2 = dict([(key, CompleteBid(value, key, Orders2Dic[key])) for key,value in ValidBids2.items()]) #二轮完整有效暗标的字典

BidResult2 = BidResult(list(chain(*CompleteBids2.values())), Teams, Database2, Budgets1)

TempSquads2 = BidResult2[0]
Database3 = BidResult2[1]
Budgets2 = BidResult2[2]

#SquadsOutput(TempSquads2, Database3, Budgets2, Teams, ManagersDic, '[-2019]二轮暗标中标名单')

def CombineSquads(squads1, squads2): #输入两个各队阵容的字典squads1和squads2，输出合并了每队阵容的字典。
    output = dict()
    for key in squads1.keys():
        squad1 = squads1[key]
        squad2 = squads2[key]
        squad = squad1 + squad2
        output[key] = sorted(squad, key = lambda entry : pos_value(entry[2]))
    return output

Squads2 = CombineSquads(Squads1, TempSquads2) #二轮暗标后阵容的字典
#SquadsOutput(Squads2, Database3, Budgets2, Teams, ManagersDic, '[-2019]二轮暗标后阵容')

###自由签#1

def CheckSign(entry, database, squads, budgets):
    announcement = ''
    team = entry[0]
    nation = entry[1]
    number = entry[2]
    key = tuple([nation, str(number)])
    position = entry[3]
    squad = squads[team]
    quad = SquadToQuad(squad)
    budget = budgets[team]
    if not key in database.keys(): # 检查球员是否在数据库中
        announcement += '无此球员 - ' + str(entry) + '\n'
        return [False, announcement]
    elif database[key]['position'] != position: # 检查球员位置是否正确
        announcement += '位置错误 - ' + str(entry) + '\n'
        return [False, announcement]
    elif database[key]['current'] != []: # 检查球员是否已被签约
        announcement += '已被签约 - ' + str(entry) + '\n'
        return [False, announcement]
    elif team in database[key]['history']: # 检查球员是否已被该队签约过
        announcement += '已签约过 - ' + str(entry) + '\n'
        return [False, announcement]
    elif sum(quad) >= 16: # 检查阵容人数是否超额
        announcement += '阵容已满 - ' + str(entry) + '\n'
        return [False, announcement]
    elif budget < 10: # 检查资金是否充足
        announcement += '资金不足 - ' + str(entry) + '\n'
        return [False, announcement]
    else:
        pos = quad[:]
        pos[pos_value(position)] = pos[pos_value(position)] + 1
        lineup = min(1, pos[0]) + min(7, pos[1] + 5, pos[1] + pos[2] + min(2, pos[3]))
        if not IsQuadUBGood(pos): # 检查各位置人数上限
            announcement += '人数超额     - ' + str(entry) + '\n'
            return [False, announcement]
        elif 10 + 10*max(0, 5 - lineup) > budget: # 检查首发阵容人数
            announcement += '首发人数不足     - ' + str(entry) + '\n'
            return [False, announcement]
        else:
            return [True, announcement]

def CompleteSign(entry): #输入entry = [team, nation, number, position], 输出阵容中格式的球员信息
    return [entry[0], tuple([entry[1],str(entry[2])]), entry[3], 17, 10, 13]

def ToTuple(string): #输入文字版的tuple，输出真正的tuple
    s = string.replace('(','')
    s = s.replace(')','')
    tup = re.split(',',s)
    nation = tup[0]
    number = tup[1]
    return (nation,number)

def PayToList(string):#输入交易一方付出的信息，输出指定格式
    output = []
    entry = re.split(' ',string)
    team = entry[0]
    output.append(team)
    info = entry[1].replace('[','')
    info = info.replace(']','')
    info = re.split(':',info)
    price = int(info[-1])
    players = list(map(ToTuple,info[:-1]))
    output.append(players)
    output.append(price)
    return output

def CheckTransfer(dic):
    entry = dic['entry']
    dictionary = dic['database']
    squads = dic['squads']
    budgets = dic['budgets']
    teams = dic['teams']
    part1 = entry[0]
    part2 = entry[1]
    team1 = part1[0]
    team2 = part2[0]
    players1 = part1[1]
    players2 = part2[1]
    price1 = part1[2]
    price2 = part2[2]
    announcement = ''
    squad1 = squads[team1]
    squad2 = squads[team2]
    newsquad1 = copy.deepcopy(squad1)
    newsquad2 = copy.deepcopy(squad2)
    budget1 = budgets[team1]
    budget2 = budgets[team2]
    newdic = copy.deepcopy(dictionary)
    for player in players1: #检查球员是否在甲队阵容&是否已经转会
        if dictionary[player]['current'] != [team1]:
            announcement = team1 + str(player) + '不在阵中 ' + str(entry)
            return [False, announcement]
        status = 0
        for profile in squad1:
            if player == profile[1]:
                status = 1
                newsquad2.append([team2] + profile[1:])
                newsquad1.remove(profile)
                newdic[player]['current'] = [team2]
                break
        if status == 0:
            announcement = team1 + '没有球员 ' + str(player) + ' ' + str(entry)
            return [False, announcement]
    for player in players2: #检查球员是否在乙队阵容
        if dictionary[player]['current'] != [team2]:
            announcement = team2 + str(player) + '不在阵中 ' + str(entry)
            return [False, announcement]
        status = 0
        for profile in squad2:
            if player == profile[1]:
                status = 1
                newsquad1.append([team1] + profile[1:])
                newsquad2.remove(profile)
                newdic[player]['current'] = [team1]
                break
        if status == 0:
            announcement = team2 + '没有球员 ' + str(player) + ' ' + str(entry)
            return [False, announcement]
    quads1 = SquadToQuad(newsquad1)
    quads2 = SquadToQuad(newsquad2)
    lineup1 = min(1,quads1[0]) + min(7, quads1[1] + 5, quads1[1] + quads1[2] + min(2, quads1[3]))
    lineup2 = min(1,quads2[0]) + min(7, quads2[1] + 5, quads2[1] + quads2[2] + min(2, quads2[3]))
    if price1 + price2 != 0: # 检查双方资金变化之和是否等于0
        announcement = '资金错误 - ' + str(entry)
        return [False, announcement]
    elif budget1 + price1 < 0: # 检查甲队资金是否充足
        announcement = team1 + '资金不足 - ' + str(entry)
        return [False, announcement]
    elif budget2 + price2 < 0: # 检查乙队资金是否充足
        announcement = team2 + '资金不足 - ' + str(entry)
        return [False, announcement]
    elif len(squad1) + len(players2) - len(players1) > 16: # 检查甲队阵容人数是否超额
        announcement = team1 + '阵容已满 - ' + str(entry)
        return [False, announcement]
    elif len(squad2) + len(players1) - len(players2) > 16: # 检查乙队阵容人数是否超额
        announcement = team2 + '阵容已满 - ' + str(entry)
        return [False, announcement]
    elif 10*max(0, 5 - lineup1) > budget1 + price1: # 检查甲队首发阵容人数
        announcement = team1 + '首发人数不足 - ' + str(entry)
        return [False, announcement]
    elif 10*max(0, 5 - lineup2) > budget2 + price2: # 检查乙队首发阵容人数
        announcement = team2 + '首发人数不足 - ' + str(entry)
        return [False, announcement]
    elif not IsQuadUBGood(quads1):
        announcement = team1 + '人数超额 - ' + str(entry)
        return [False, announcement]
    elif not IsQuadUBGood(quads2):
        announcement = team2 + '人数超额 - ' + str(entry)
    else:
        squads[team1] = newsquad1
        squads[team2] = newsquad2
        budgets[team1] = budget1 + price1
        budgets[team2] = budget2 + price2
        return [True, {'database':newdic, 'squads':squads, 'budgets':budgets}]

def WindowToList(text):
    output = []
    while 1:
        line = text.readline()
        if not line:
            return output
        else:
            if ';' in line:
                entry = re.split(';',line)
                if len(entry) != 2:
                    print(entry)
                else:
                    transfer = list(map(PayToList,entry))
                    output.append(['t'] + [transfer])
            else:
                line = line.rstrip('\n')
                entry = re.split('[\s]+',line)
                output.append(['s'] + [entry])
    return output

def Window(dic):
    text = dic['text']
    database = copy.deepcopy(dic['database'])
    squads = dic['squads']
    budgets = dic['budgets']
    teams = dic['teams']
    managers = dic['managers']
    filename = dic['filename']
    isprint = dic['print']
    windowlist = WindowToList(text)
    errorinfo = ''
    for record in windowlist:
        recordtype = record[0]
        if recordtype == 's':
            entry = record[1]
            status = entry[-1]
            if status == 's':
                check = CheckSign(entry, database, squads, budgets)
                errorinfo += check[1]
                if check[0]:
                    team = entry[0]
                    nation = entry[1]
                    number = entry[2]
                    key = tuple([nation, str(number)])
                    position = entry[3]
                    sign = CompleteSign(entry)
                    database[key]['current'] = [team] #球员主页current信息添加新签约的球队
                    database[key]['history'].append(team) #球员主页history信息添加新签约的球队
                    squads[team].append(sign) #阵容添加新签约的球员
                    budgets[team] -= 10 #玩家资金减10
            elif status == 'f':
                    team = entry[0]
                    nation = entry[1]
                    number = entry[2]
                    key = tuple([nation, str(number)])
                    position = database[key]['position']
                    squad = squads[team][:]
                    pos_value = pos_value(position)
                    for player in squad:
                        if player[1] == key and player[2] == position:
                            squad.remove(player)  #阵容删去新解约的球员
                            squads[team] = squad
                            database[key]['current'] = [] #球员主页current信息删去新解约的球队
        elif recordtype == 't':
            transfer = record[1]
            check = CheckTransfer({'entry':transfer, 'database':database, 'squads':squads, 'budgets':budgets, 'teams':teams})
            if not check[0]:
                errorinfo += check[1]
            else:
                database = check[1]['database']
                squads = check[1]['squads']
                budgets = check[1]['budgets']
    for team in teams:
        squads[team] = sorted(squads[team], key = lambda entry: pos_value(entry[2]))
    # 打印错误信息
    if errorinfo != '':
        print(errorinfo)
    # 输出转会窗后结果
    if isprint:
        SquadsOutput(squads, database, budgets, teams, managers, filename)
    return {'squads':squads, 'database':database, 'budgets':budgets}

Window1 = open("window1.txt")

Window1dic = {'text':Window1, 'database':Database3, 'squads':Squads2, 'budgets': Budgets2, 'teams':Teams, 'managers':ManagersDic, 'filename':'[-2019]自由签#1后阵容', 'print':False}

WindowOutput1 = Window(Window1dic)

Squads3 = WindowOutput1['squads']
Database4 = WindowOutput1['database']
Budgets3 = WindowOutput1['budgets']

### 8强赛
# 晋级国家共8个
Nations = ['Argentina', 'Brazil', 'Chile', 'Colombia', 'Paraguay', 'Peru', 'Uruguay', 'Venezuela']

# 游戏晋级球队，共8支
Teams = ['ARG','BRA','CHL','COL','JPN','PER','PRY','QAT']

# 各队主教练ID，顺序与Teams对应。
Managers = ['Jasper', 'Montella', 'solojuve', 'KakaHiguain', 'zincum & jiamingpku', 'rhyshm', 'flybut', 'angelqi']
ManagersDic = dictize(Teams, Managers) #玩家信息的字典

# 更新大名单和资金

Budgets4 = [81,221,103,134,71,94,113,103]
Budgets4 = dictize(Teams, Budgets4)  #资金信息的字典

Database5 = copy.deepcopy(Database4)

Database5 = {k:v for k,v in Database4.items() if k[0] in Nations} #去掉未晋级国家的球员
for key in Database5.keys():
    if Database5[key]['current'] != []:
        if not Database5[key]['current'][0] in Teams:
            Database5[key]['current'] = [] #current信息去掉未晋级的球队

Squads4 = {k:v for k,v in Squads3.items() if k in Teams} #去掉未晋级的球队的阵容

for key in Squads4.keys():
    squad = []
    for entry in Squads4[key]:
        if entry[1][0] in Nations:
            squad.append(entry)
    Squads4[key] = squad #更新晋级的球队的阵容

#SquadsOutput(Squads4, Database5, Budgets4, Teams, ManagersDic, '[-2019]8强赛暗标前阵容')

### 竞标过程

Orders3 = [4,7,2,1,6,8,5,3] #8强赛暗标各队投标顺序
Orders3Dic = dictize(Teams, Orders3)

Bids3 = BidsDic(Teams, Nations, Black1Dic, 3) #8强赛暗标的列表形式标书的字典
ValidBids3 =dict([(key, CheckBid(key, value, Database5, SquadToQuad(Squads4[key]), Budgets4[key])[0]) for key,value in Bids3.items()]) #8强赛有效暗标的字典
CompleteBids3 = dict([(key, CompleteBid(value, key, Orders3Dic[key])) for key,value in ValidBids3.items()]) #8强赛完整有效暗标的字典

BidResult3 = BidResult(list(chain(*CompleteBids3.values())), Teams, Database5, Budgets4)

TempSquads3 = BidResult3[0]
Database6 = BidResult3[1]
Budgets5 = BidResult3[2]

#SquadsOutput(TempSquads3, Database6, Budgets5, Teams, ManagersDic, '[-2019]8强赛暗标中标名单')

Squads5 = CombineSquads(Squads4, TempSquads3) #8强赛暗标后阵容的字典
#SquadsOutput(Squads5, Database6, Budgets5, Teams, ManagersDic, '[-2019]8强赛暗标后阵容')


### 转会窗#2

Window2 = open("window2.txt")

Window2dic = {'text':Window2, 'database':Database6, 'squads':Squads5, 'budgets': Budgets5, 'teams':Teams, 'managers':ManagersDic, 'filename':'[-2019]8强赛自由签后阵容', 'print':False}

WindowOutput2 = Window(Window2dic)

Squads6 = WindowOutput2['squads']
Database7 = WindowOutput2['database']
Budgets6 = WindowOutput2['budgets']

### 4强赛
# 晋级国家共4个
Nations = ['Argentina', 'Brazil', 'Chile', 'Peru']

# 游戏晋级球队，共4支
Teams = ['ARG','COL','PRY','QAT']

# 各队主教练ID，顺序与Teams对应。
Managers = ['Jasper',  'KakaHiguain', 'flybut', 'angelqi']
ManagersDic = dictize(Teams, Managers) #玩家信息的字典

# 更新大名单和资金

Budgets7 = [71,108,83,82]
Budgets7 = dictize(Teams, Budgets7)  #资金信息的字典

Database8 = copy.deepcopy(Database7)

Database8 = {k:v for k,v in Database8.items() if k[0] in Nations} #去掉未晋级国家的球员
for key in Database8.keys():
    if Database8[key]['current'] != []:
        if not Database8[key]['current'][0] in Teams:
            Database8[key]['current'] = [] #current信息去掉未晋级的球队

Squads7 = {k:v for k,v in Squads6.items() if k in Teams} #去掉未晋级的球队的阵容

for key in Squads7.keys():
    squad = []
    for entry in Squads7[key]:
        if entry[1][0] in Nations:
            squad.append(entry)
    Squads7[key] = squad #更新晋级的球队的阵容

pick1 = open("pick1.txt")

pick1dic = {'text':pick1, 'database':Database8, 'squads':Squads7, 'budgets': Budgets7, 'teams':Teams, 'managers':ManagersDic, 'filename':'[-2019]4强赛挑人后阵容', 'print':False}

pick1Output = Window(pick1dic)

Squads8 = pick1Output['squads']
Database9 = pick1Output['database']
Budgets8 = [71,108,83,82]
Budgets8 = dictize(Teams, Budgets8)  #资金信息的字典

#SquadsOutput(Squads8, Database9, Budgets8, Teams, ManagersDic, '[-2019]4强赛暗标前阵容')

### 竞标过程

Orders4 = [3,2,4,1] #4强赛暗标各队投标顺序
Orders4Dic = dictize(Teams, Orders4)

Bids4 = BidsDic(Teams, Nations, Black1Dic, 4) #4强赛暗标的列表形式标书的字典
ValidBids4 =dict([(key, CheckBid(key, value, Database9, SquadToQuad(Squads8[key]), Budgets8[key])[0]) for key,value in Bids4.items()]) #4强赛有效暗标的字典
CompleteBids4 = dict([(key, CompleteBid(value, key, Orders4Dic[key])) for key,value in ValidBids4.items()]) #4强赛完整有效暗标的字典

BidResult4 = BidResult(list(chain(*CompleteBids4.values())), Teams, Database9, Budgets8)

TempSquads4 = BidResult4[0]
Database10 = BidResult4[1]
Budgets9 = BidResult4[2]

#SquadsOutput(TempSquads4, Database10, Budgets9, Teams, ManagersDic, '[-2019]4强赛暗标中标名单')

Squads9 = CombineSquads(Squads8, TempSquads4) #4强赛暗标后阵容的字典
#SquadsOutput(Squads9, Database10, Budgets9, Teams, ManagersDic, '[-2019]4强赛暗标后阵容')

### 转会窗#3

Window3 = open("window3.txt")

Window3dic = {'text':Window3, 'database':Database10, 'squads':Squads9, 'budgets': Budgets9, 'teams':Teams, 'managers':ManagersDic, 'filename':'[-2019]4强赛自由签后阵容', 'print':False}

WindowOutput3 = Window(Window3dic)

Squads10 = WindowOutput3['squads']
Database11 = WindowOutput3['database']
Budgets10 = WindowOutput3['budgets']

### 决赛

# 游戏晋级球队，共2支
Teams = ['COL','QAT']

# 各队主教练ID，顺序与Teams对应。
Managers = ['KakaHiguain', 'angelqi']
ManagersDic = dictize(Teams, Managers) #玩家信息的字典

# 更新大名单和资金

Budgets11 = [106,59]
Budgets11 = dictize(Teams, Budgets11)  #资金信息的字典

Database12 = copy.deepcopy(Database11)

for key in Database12.keys():
    if Database12[key]['current'] != []:
        if not Database12[key]['current'][0] in Teams:
            Database12[key]['current'] = [] #current信息去掉未晋级的球队

Squads11 = {k:v for k,v in Squads10.items() if k in Teams} #去掉未晋级的球队的阵容

pick2 = open("pick2.txt")

pick2dic = {'text':pick2, 'database':Database12, 'squads':Squads11, 'budgets': Budgets11, 'teams':Teams, 'managers':ManagersDic, 'filename':'[-2019]决赛挑人后阵容', 'print':False}

pick2Output = Window(pick2dic)

Squads12 = pick2Output['squads']
Database13 = pick2Output['database']
Budgets12 = [106,59]
Budgets12 = dictize(Teams, Budgets12)  #资金信息的字典

#SquadsOutput(Squads12, Database13, Budgets12, Teams, ManagersDic, '[-2019]决赛暗标前阵容')

### 竞标过程

Orders5 = [1,2] #决赛暗标各队投标顺序
Orders5Dic = dictize(Teams, Orders5)

Bids5 = BidsDic(Teams, Nations, Black1Dic, 5) #决赛暗标的列表形式标书的字典
ValidBids5 =dict([(key, CheckBid(key, value, Database13, SquadToQuad(Squads12[key]), Budgets12[key])[0]) for key,value in Bids5.items()]) #决赛有效暗标的字典
CompleteBids5 = dict([(key, CompleteBid(value, key, Orders5Dic[key])) for key,value in ValidBids5.items()]) #决赛完整有效暗标的字典

BidResult5 = BidResult(list(chain(*CompleteBids5.values())), Teams, Database13, Budgets12)

TempSquads5 = BidResult5[0]
Database14 = BidResult5[1]
Budgets13 = BidResult5[2]

SquadsOutput(TempSquads5, Database14, Budgets13, Teams, ManagersDic, '[-2019]决赛暗标中标名单')

Squads13 = CombineSquads(Squads12, TempSquads5) #决赛暗标后阵容的字典
SquadsOutput(Squads13, Database14, Budgets13, Teams, ManagersDic, '[-2019]决赛暗标后阵容')
