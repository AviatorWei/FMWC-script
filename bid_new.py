# 北大未名BBS体育游戏版FM-America 2019游戏
# BY survivor@BDWM
# Python 3.7

import re
import pickle
import string
import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import chain

###基本信息

# 2022世界杯参赛国家，共32个
Nations = ['Ecuador', 'Netherlands', 'Qatar', 'Senegal', 'England', 'Iran', 'USA', 'Wales',
           'Argentina', 'Mexico', 'Poland', 'SaudiArabia', 'Australia', 'Denmark', 'France', 'Tunisia',
           'CostaRica', 'Germany', 'Japan', 'Spain', 'Belgium', 'Canada', 'Croatia', 'Morocco',
           'Brazil', 'Cameroon', 'Serbia', 'Switzerland', 'Ghana', 'Portugal', 'SouthKorea', 'Uruguay']

# FMWC游戏参赛球队，共24支
Participants = [
    ['Sang', 'ESP', 10, 9], ['angelqi', 'BRA', 9, 11], ['augustusc', 'MEX', 8, 10], ['aidengazhaer', 'NED', 24, 23],
    ['since', 'QAT', 22, 15], ['KakaHiguain', 'MAR', 13, 16], ['Nocchiere', 'DEN', 18, 20], ['sixingdeguo', 'SRB', 12, 19],
    ['Arjen', 'SUI', 3, 5], ['Zeymax', 'ARG', 14, 14], ['RealMadrid', 'POR', 19, 17], ['rhyshm', 'FRA', 23, 7],
    ['HydraliskIII', 'KSA', 11, 4], ['solojuve', 'ECU', 21, 22], ['Montella', 'CAN', 7, 12], ['pkuarsene', 'GHA', 20, 18],
    ['weilovebvb', 'CRC', 17, 8], ['twa', 'WAL', 16, 200], ['linsage', 'SEN', 15, 21], ['IanWalls', 'ENG', 200, 13],
    ['jiamingpku', 'URU', 4, 3], ['Augustus', 'BEL', 100, 100], ['Jimmywiki', 'JPN', 5, 24], ['survivor', 'USA', 6, 6],
]

Teams = [x[1] for x in Participants]
# 各队主教练ID，顺序与Teams对应。
Managers = [x[0] for x in Participants]

original_roster = 'FMWC2022大名单.xlsx'
new_roster = 'FMWC2022大名单-3new.xlsx'

### 人数要求
SquadUB = 15  # 阵容人数上限
positions = ['G', 'D', 'M', 'F']
PosUB = [2, 4, 6, 3]  # 阵容各位置人数上限
LineupLB = 6  # 首发人数下限
LineupPosUB = [1, 2, 3, 2]  # 首发各位置人数上限

Budget0 = 600  # 各队初始资金数额
Orders1 = [x[2] for x in Participants]
Orders2 = [x[3] for x in Participants]
cur_round = 2

##############################################################################

def dictize(keys, values):  # 输入键列表keys和值列表values，输出对应的字典dic。
    assert len(keys) == len(values), f"Length not equal {keys}, {values}"
    dic = {}
    for key, value in zip(keys, values):
        assert key not in dic, f"Key existed {key}-{value}, original {dic[key]}"
        dic[key] = value
    return dic


def str_len(string):  # 输入字符串string，输出string的字符个数，适配中文。
    try:
        row_l = len(string)
        utf8_l = len(string.encode('utf-8'))
        return int((utf8_l - row_l) / 2 + row_l)
    except:
        return None


def pos_value(p):  # 输入位置的单个字符，输出数值0，1，2，3.
    pos_dict = dict([(v, i) for i, v in enumerate(positions)])
    return pos_dict[p]


def SquadToQuad(List):  # 输入一个多名球员阵容List，输出其四个位置上球员个数的列表。
    output = [0] * 4
    for entry in List:
        pos = entry[2]
        output[pos_value(pos)] += 1
    return output


def BuildDatabase(worksheet):
    output = dict()
    for row in worksheet.iter_rows(min_row=2, max_col=4):
        entry = [x.value for x in row]
        key = tuple([entry[2], str(entry[3])])
        value = dict()
        value['name'] = entry[0].strip(' ')  # 球员姓名
        value['position'] = entry[1].strip(' ')  # 球员位置
        value['current'] = []  # 球员当前所属玩家列表
        value['history'] = []  # 球员属于过的全部玩家列表
        value['price'] = None
        output[key] = value
    return output


def PosQuad(bid):  # 输入列表形式的标书bid，输出标书中四个位置的球员个数的列表。
    output = [0, 0, 0, 0]
    for entry in bid:
        position = entry[1]
        output[pos_value(position)] += 1
    return output


def SquadPosCnt(squad):
    output = [0, 0, 0, 0]
    for entry in squad:
        position = entry[2]
        output[pos_value(position)] += 1
    return output


def IsQuadUBGood(quad):  # 输入阵容四个位置人数列表quad，输出其是否满足阵容人数上限要求。
    if sum(quad) > SquadUB:
        return False
    elif quad[0] > PosUB[0]:
        return False
    else:
        for index in range(1, 4):
            if sum(quad[index:]) > sum(PosUB[index:]):
                return False
        return True


def NeededPlayer(bid, currentQuad):  # 输入标书bid和该玩家阵容中四个位置已有的球员个数列表currentQuad，输出在暗标全部命中的情况下，该玩家阵容需要补充的球员总数。
    quad = [PosQuad(bid)[index] + currentQuad[index] for index in range(4)]
    minNonGK = LineupLB - 1
    maxForward = LineupPosUB[-1]
    lineup = min(1, quad[0]) \
             + min(minNonGK, quad[1] + minNonGK,
                   quad[1] + quad[2] + min(maxForward, quad[3]))
    # 第一项是标书在门将位置对首发的最大贡献，第二项是标书在后卫、中场、前锋位置对首发的最大贡献之和。
    return max(0, 6 - lineup)


# 读取xlsx标书
def read_bid(root, name, round):
    path = root + '/' + name.upper() + str(round) + '.xlsx'
    bid = load_workbook(path)
    sheet = bid.active
    assert sheet is not None, f"sheet nonetype {path}"
    output = []
    for row in sheet.iter_rows(min_row=2, max_col=6):
        entry = [x.value for x in row]
        if entry[1] is not None:
            key = tuple([entry[4], str(entry[5])])  # key=(球员国家, 号码)
            name = entry[2].strip(' ')  # 球员姓名
            pos = entry[3].strip(' ')  # 球员位置
            order = 99 if entry[0] is None else entry[0]  # 排名
            price = entry[1]  # 出价
            assert isinstance(order, int), f'{path} order not int: {order}'
            assert isinstance(price, int), f'{path} price not int: {price}'
            bid = [key, pos, order, price]
            output.append(bid)
    return output


def BidsDic(root, teams, nations, blackdic, rd):  # 输入玩家信息teams，球队信息nations，不能签约球员信息的字典blackdic，输出各队列表形式标书的字典。
    bids = dict()
    for team in teams:
        # black = blackdic[team]
        bid = read_bid(root, team, rd)
        bids[team] = bid
    return bids


### 列表标书处理
def TopPlayer(bid, pos=''):  # 输入列表形式的标书bid，输出该标书中出价最高的一个投标。出价相同时，再依次比较位置, 顺位。
    if bid == []:  # 如果输入为空，返回空列表。
        return []
    else:  # 如果输入不为空
        bidList = bid.copy()
        if pos != '':
            bidList = [x for x in bid if x[1] == pos]
        sortedlist = sorted(bidList, key=lambda x: (-x[3], pos_value(x[1]), x[2]))  # 依次比较：-出价, 位置数值, 顺位，都是数值小的顺序优先。
        return sortedlist[0]  # 返回出价最高的投标


def Budget(bid):  # 输入列表形式的标书bid，输出标书总出价。
    return sum([entry[-1] for entry in bid])


# 输入玩家Team，其列表形式的标书List，球员数据库dictionary，该玩家当前阵容currentQuad和当前资金budget，检查标书合法性并返回[列表形式的所有有效投标bid,无效投标的报错信息文本announcement]。
def CheckBid(team, oriBid, db, currentQuad, budget):
    bid = []
    announcement = ''
    for entry in oriBid:
        key = entry[0]
        position = entry[1]
        price = entry[3]
        if not key in db.keys():  # 检查球员是否在数据库中
            announcement = announcement + '无此球员     - ' + str(entry)
        elif db[key]['current'] != []:  # 检查球员是否已经被签约
            announcement = announcement + '已被签约     - ' + str(entry)
        elif team in db[key]['history']:  # 检查球员是否曾经被team签约
            announcement = announcement + '无资格签约    - ' + str(entry)
        elif db[key]['position'] != position:  # 检查球员的位置是否正确
            announcement = announcement + '位置错误     - ' + str(entry)
        elif price < 10:  # 检查出价是否至少为10
            announcement = announcement + '出价小于10   - ' + str(entry)
        else:  # 正确则放进bid列表
            bid.append(entry)
    # while len(bid) + sum(currentQuad) > SquadUB:  # 标书多于SquadUB人时，移除最高价球员直到剩SquadUB人
    #     top = TopPlayer(bid)
    #     announcement = announcement + '人数超额 - ' + (len(bid) + sum(currentQuad))
    #     bid.remove(top)
    curCnt = 0
    for posIdx in range(3, -1, -1):
        while PosQuad(bid)[posIdx] + currentQuad[posIdx] > sum(PosUB[posIdx:]) - curCnt: # 位置人数超额，不考虑总体人数
            top = TopPlayer(bid, positions[posIdx])
            announcement = announcement + f'{positions[posIdx]}人数超标 - ' + str(top)
            bid.remove(top)
        curCnt += PosQuad(bid)[posIdx]
    while Budget(bid) > budget and len(bid) > 0:  # 标书总出价大于剩余资金budget时，移除最高价球员直到总出价不大于budget
        top = TopPlayer(bid)
        announcement = announcement + '预算超额 - ' + str(Budget(bid)) + str(top)
        bid.remove(top)
    while Budget(bid) + 10 * NeededPlayer(bid, currentQuad) > budget and len(bid) > 0:  # 标书总出价使得全中可能导致凑不齐首发阵容人数下限，移除最高价球员直到满足要求
        top = TopPlayer(bid)
        announcement = announcement + '首发人数不足'
        bid.remove(top)

    # 去掉无效投标后，更新顺位
    # if bid == []:
    #     return [bid, announcement]
    # else:
    #     output = []
    #     currentpos = bid[0][1] #标书首行的位置
    #     counter = 0
    #     for entry in bid:
    #         pos = entry[0]
    #         if pos != currentpos:
    #             counter = 1
    #             currentpos = pos
    #         else:
    #             counter = counter + 1
    #         output.append([entry[0], entry[1], counter, entry[3]])
    output = bid
    if announcement != '':
        print(team, announcement)
    return [output, announcement]


def CompleteBid(bid, team, order):  # 输入同一玩家列表形式的所有有效投标bid, 每条投标添加队名和投标顺序成为完整标书 --  [玩家, (国家, 号码), 位置, 顺位, 出价, 投标顺序]
    return [[team] + entry + [order] for entry in bid]


### 竞标过程

def BidCompare(entries):  # 输入对同一球员的多笔有效投标的列表entries，输出中标的投标。
    if len(entries) == 1:
        return entries[0], entries
    else:
        ordered = sorted(entries,
                         key=lambda entry: (-entry[4], entry[3], entry[5]))  # 依次按照出价（从大到小）, 顺位（从小到大）, 投标顺序（从小到大）进行排序。
        return ordered[0], ordered


def BidResult(bids, teams, dictionary,
              budgets):  # 输入列表形式的所有有效投标bids, 玩家teams，数据库dictionary和预算字典budgets，比较全部有效标书的投标，得出中标情况。
    players = dictionary.keys()  # 所有的球员列表
    profiles = dict.fromkeys(players, [])  # 创建每个球员全部被投标信息的字典，球员是键，投标信息的列表是值。
    newdictionary = copy.deepcopy(dictionary)
    newbudgets = budgets.copy()
    output = dict()
    for team in teams:
        output[team] = []
    for bid in bids:
        key = bid[1]
        profiles[key] = profiles[key].copy() + [bid]  # 更新字典profile的值
    allBids = []
    for player in profiles.keys():
        if profiles[player] == []:  # 如果球员player不出现在暗标中，则直接跳过。
            continue
        else:  # 如果球员player出现在暗标中
            successbid, orderedBid = BidCompare(profiles[player])  # 得到对球员player中标的投标
            team = successbid[0]
            price = successbid[4]
            output[team].append(successbid)  # 添加到输出
            profile = dictionary[player]  # 当前数据库中球员player的信息
            profile['current'] = [team]  # 添加其当前所属球队的信息
            profile['price'] = price
            profile['history'].append(team)  # 添加其曾经所属球队的信息
            newdictionary[player] = profile  # 更新数据库中球员player的信息
            newbudgets[team] = newbudgets[team] - price  # 更新玩家team的资金信息
            allBids.append(orderedBid)
    for team in teams:
        output[team] = sorted(output[team], key=lambda entry: pos_value(entry[2]))
    return [output, newdictionary, newbudgets, allBids]  # 输出各队中标信息output，新数据库newdictionary和新资金信息newbudgets


### 输出文本处理
def LineToTxt(parts):  # 输入单行格式的参数parts = [part1, part2, ...]，输出按照此格式的文本。其中每个part = [内容, 长度, 对齐方式]
    output = ''
    for index in range(len(parts)):
        part = parts[index]
        string = str(part[0])
        length = part[1]
        align = part[2]
        if align == 'r':
            output = output + ' ' * (max(length - str_len(string), 0)) + string
        elif index == len(parts) - 1:
            output = output + string
        else:
            output = output + string + ' ' * (max(length - str_len(string), 0))
    output = output + '\n'
    return output


def SquadToText(squad, database, budget, team,
                manager):  # 输入一个玩家team的阵容squad，以及其资金budget和大师manager，输出该玩家阵容按指定格式的文本output。
    size = len(squad)  # 阵容人数
    output = team + ' ' + str(size) + '人' + ' ' * (33 - len(team) - len(str(size))) + '剩余资金' + ' ' * (
            4 - len(str(budget))) + str(budget) + 'm\n'
    output = output + manager + '\n'
    for entry in squad:
        nation = entry[1][0]
        number = entry[1][1] + '号'
        name = database[entry[1]]['name']
        position = entry[2]
        price = str(entry[4]) + 'm'
        info = [team, name, nation, number, position, price]
        card = len(info)
        lengths = [5, 20, 12, 8, 4, 6] # Originally 18 for [1], change to 20 because of SMS
        aligns = ['l', 'l', 'l', 'r', 'r', 'r']
        parts = [[string, length, align] for string, length, align in zip(info, lengths, aligns)]
        output += LineToTxt(parts)
    output += '\n'
    return output


def SquadsOutput(squads, database, budgets, teams, managers,
                 filename):  # 输入各玩家的字典teams，其阵容的字典squads，资金的字典budgets，大师的字典managers，以及输出文件的文件名filename，输出各玩家阵容按指定格式的文本output并保存到filename.txt
    output = ''
    for key in teams:
        squad = squads[key]
        budget = budgets[key]
        manager = managers[key]
        output += SquadToText(squad, database, budget, key, manager)
    file = open(filename, 'w')
    file.write(output)
    file.close()
    return output


def BidToText(bid, database):
    team = bid[0]
    key = bid[1]
    nation, number = key
    number = number + '号'
    name = database[key]['name']
    position = bid[2]
    order = bid[3]
    price = str(bid[4]) + 'm'
    info = [order, price, name, position, nation, number, team]
    lengths = [4, 5, 20, 3, 12, 4, 4] # Was orignially 18 for [2], changed to 20 because of SMS
    aligns = ['l', 'l', 'l', 'l', 'l', 'l', 'r']
    parts = [[string, length, align] for string, length, align in zip(info, lengths, aligns)]
    return LineToTxt(parts)


def BidsOutput(bids, database, teams, filename):
    output = ''
    for bid in bids:
        output += ''.join([BidToText(entry, database) for entry in bid])
        output += '\n'
    file = open(filename, 'w')
    file.write(output)
    file.close()
    return output


###二轮暗标
def UpdateDB(workbook, database):  # 输入excel工作簿workbook，工作表名title和旧数据库database，输出新数据库。
    output = copy.deepcopy(database)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, max_col=6):
        entry = [x.value for x in row]
        key = tuple([entry[2], str(entry[3])])
        if key not in database.keys():
            value = dict()
            value['name'] = entry[0].strip(' ')  # 球员姓名
            value['position'] = entry[1].strip(' ')  # 球员位置
            value['current'] = []  # 球员当前所属玩家列表
            value['history'] = []  # 球员属于过的全部玩家列表
            value['price'] = None
        elif database[key]['current'] != []:
            team = database[key]['current'][0]
            price = database[key]['price']
            team_cell = row[4]
            team_cell.value = team
            price_cell = row[5]
            price_cell.value = price
    workbook.save(new_roster)
    return output


if __name__ == "__main__":
    ManagersDic = dictize(Teams, Managers)  # 玩家信息的字典
    Budgets0Dic = dictize(Teams, [Budget0] * len(Teams))  # 资金信息的字典

    OrdersDic = []
    OrdersDic.append(dictize(Teams, Orders1))
    OrdersDic.append(dictize(Teams, Orders2))

    Black1Dic = dictize(Teams, [[]] * len(Teams))
    SquadDict = dictize(Teams, [[]] * len(Teams))

    ### 建立大名单的数据库
    Roster1 = load_workbook(original_roster)
    Worksheet1 = Roster1.active
    # 存储当前数据库
    # if cur_round == 1:
    Database = BuildDatabase(Worksheet1)
    # else:
    #     # 从pickle文件加载已有的数据库
    #     with open(f'FMWC-2022-Database-{cur_round}.pickle', 'rb') as file:
    #         Database = pickle.load(file)
    for round in range(1, cur_round + 1):
        root = f'../bids-{round}'
        Bids1 = BidsDic(root, Teams, Nations, Black1Dic, round)  # 一轮暗标的列表形式标书的字典

        ValidBids1 = dict([(team, CheckBid(team, bid, Database, SquadPosCnt(SquadDict[team]), Budgets0Dic[team])[0]) for team, bid in
                           Bids1.items()])  # 一轮有效暗标的字典
        CompleteBids1 = dict(
            [(team, CompleteBid(bid, team, OrdersDic[round - 1][team])) for team, bid in ValidBids1.items()])  # 一轮完整有效暗标的字典

        result = BidResult(list(chain(*CompleteBids1.values())), Teams, Database, Budgets0Dic)
        Squads1 = result[0]
        Database1 = result[1]
        Budgets1 = result[2]
        Bids1 = result[3]

        for team in SquadDict.keys():
            SquadDict[team] = SquadDict[team] + Squads1[team]
            SquadDict[team] = sorted(SquadDict[team], key=lambda entry: pos_value(entry[2]))
        Database = Database1
        Budgets0Dic = Budgets1
        print(SquadDict)
        # print(Budgets0Dic)

    BidsOutput(Bids1, Database, Teams, '../暗标公示.txt')
    SquadsOutput(SquadDict, Database, Budgets0Dic, Teams, ManagersDic, '../暗标后阵容.txt')

    Roster2 = load_workbook(new_roster)  # 二轮大名单
    Database2 = UpdateDB(Roster2, Database)  # 次轮暗标数据库

    with open(f'FMWC-2022-Database-{cur_round + 1}.pickle', 'wb') as handle:
        pickle.dump(Database, handle, protocol=pickle.HIGHEST_PROTOCOL)
